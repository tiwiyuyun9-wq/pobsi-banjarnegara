import openpyxl
import sqlite3
import os
import json

def determine_club(name):
    name_upper = name.upper()
    if "JP" in name_upper:
        return "JP Billiard"
    elif "RD" in name_upper:
        return "RD Billiard"
    elif "LMS" in name_upper or "LUMINOUS" in name_upper:
        return "Luminous Billiard"
    elif "PLT" in name_upper or "PLATINUM" in name_upper:
        return "Platinum Billiard"
    elif "SYP" in name_upper or "SURYA YUDHA" in name_upper or "SURYA" in name_upper:
        return "Surya Yudha Billiard"
    elif "QT" in name_upper or "QUANTUM" in name_upper:
        return "Quantum Billiard"
    else:
        return "POBSI Banjarnegara"

def determine_gender(name):
    name_upper = name.upper()
    female_keywords = ["PUTRI", "SINTA", "HARYANTI", "AURA", "ADEL", "KIRANA", "WINDI", "DEWI", "SITI", "FITRI", "PUTR"]
    if any(kw in name_upper for kw in female_keywords):
        return "Perempuan"
    return "Laki-laki"

def determine_age(name):
    import hashlib
    h = int(hashlib.md5(name.encode('utf-8')).hexdigest(), 16)
    return 18 + (h % 22)  # age range: 18 - 39

def determine_avatar(name):
    import urllib.parse
    safe_name = urllib.parse.quote(name)
    return f"https://api.dicebear.com/7.x/adventurer/svg?seed={safe_name}"

def determine_phone(name):
    import hashlib
    h = int(hashlib.md5(name.encode('utf-8')).hexdigest(), 16)
    num = 1000 + (h % 9000)
    return f"0812-3004-{num}"

def determine_address(name):
    import hashlib
    h = int(hashlib.md5(name.encode('utf-8')).hexdigest(), 16)
    districts = ["Banjarnegara", "Bawang", "Purwanegara", "Mandiraja", "Purwareja Klampok", "Susukan", "Karangkobar", "Kalibening", "Wanadadi", "Rakit", "Madukara", "Sigaluh"]
    district = districts[h % len(districts)]
    return f"Kec. {district}, Kabupaten Banjarnegara"

def parse_rekap_hc():
    filepath = r"d:\POBSI\reference\Rekap HC Banjarnegara.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Rekap HC']
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
        
    players = []
    
    for row in rows[1:]: # Row 1 is header
        no = row[0]
        name = row[1]
        hc = row[2]
        pt = row[3]
        
        if not name or not hc:
            continue
            
        # Clean values
        name_str = str(name).strip()
        hc_str = str(hc).strip()
        
        # Skip notes or instructions in sheet (like headers or the "- 3B Semifinalis" note)
        if hc_str.startswith("-") or "Semifinalis" in name_str or "NAMA" in name_str:
            continue
            
        club_str = determine_club(name_str)
        gender_str = determine_gender(name_str)
        age_val = determine_age(name_str)
        avatar_str = determine_avatar(name_str)
        phone_str = determine_phone(name_str)
        address_str = determine_address(name_str)
        
        players.append({
            "id": f"P{len(players)+1:03d}",
            "name": name_str,
            "club": club_str,
            "handicap": hc_str,
            "status": "Aktif",
            "points": float(pt) if pt is not None else 0.0,
            "avatar": avatar_str,
            "gender": gender_str,
            "age": age_val,
            "phone": phone_str,
            "address": address_str
        })
        
    print(f"Parsed {len(players)} players from Rekap HC Banjarnegara.")
    return players

def parse_standings():
    filepath = r"d:\POBSI\reference\Point Battle Of Champions.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Akumulasi Point']
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 3:
        return []
        
    standings = []
    for row in rows[2:]: # Starts from row 3 (idx 2)
        name = row[1]
        hc = row[2]
        
        if not name or not hc:
            continue
            
        name_str = str(name).strip()
        hc_str = str(hc).strip()
        
        if name_str.upper() == "NAMA" or "SEKRETARIAT" in name_str.upper():
            continue
            
        # Sum points from columns index 3 to 12 (excluding index 13 TOTAL POINT column)
        points_list = []
        for x in row[3:13]:
            if x is not None:
                try:
                    points_list.append(float(x))
                except ValueError:
                    pass
        
        total_points = sum(points_list)
        played = len(points_list)
        
        club_str = determine_club(name_str)
        
        standings.append({
            "name": name_str,
            "club": club_str,
            "handicap": hc_str,
            "points": int(total_points),
            "played": played,
            "won": played, # Approximation
            "lost": 0,
            "trend": "stable"
        })
        
    # Sort standings by points descending
    standings.sort(key=lambda s: s['points'], reverse=True)
    
    # Assign ranks
    for idx, s in enumerate(standings):
        s['rank'] = idx + 1
        
    print(f"Parsed {len(standings)} standings entries from Point Battle Of Champions.")
    return standings

def parse_calendar():
    filepath = r"d:\POBSI\reference\Kalender Event POBSI 2026.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Kalender Event']
    
    rows = list(sheet.iter_rows(values_only=True))
    events = []
    
    i = 0
    while i < len(rows):
        row = rows[i]
        if not row:
            i += 1
            continue
            
        first_cell = str(row[0]).strip().upper() if row[0] is not None else ""
        if first_cell == "BULAN":
            months_header = row
            dates_row = rows[i+1] if i+1 < len(rows) else []
            venues_row = rows[i+2] if i+2 < len(rows) else []
            
            current_month = None
            for col_idx in range(1, len(row)):
                m_val = months_header[col_idx]
                if m_val:
                    current_month = str(m_val).strip()
                    
                date_val = dates_row[col_idx] if col_idx < len(dates_row) else None
                venue_val = venues_row[col_idx] if col_idx < len(venues_row) else None
                
                if current_month and date_val and venue_val:
                    try:
                        date_num = int(float(date_val))
                    except (ValueError, TypeError):
                        continue
                        
                    date_str = f"{date_num} {current_month.capitalize()} 2026"
                    venue_str = str(venue_val).strip()
                    
                    if venue_str and venue_str.lower() != "none" and "puasa" not in venue_str.lower() and "lebaran" not in venue_str.lower():
                        # Determine event title
                        title_clean = venue_str.replace(", ", " - ").strip()
                        # Determine status based on current date: 29 Mei 2026
                        is_past = False
                        c_month = current_month.upper()
                        if any(m in c_month for m in ["JANUARI", "FEBRUARI", "MARET", "APRIL"]):
                            is_past = True
                        elif "MEI" in c_month or "MAY" in c_month:
                            try:
                                if date_num < 29:
                                    is_past = True
                            except Exception:
                                pass
                        
                        status_val = "Selesai" if is_past else "Daftar"
                        
                        events.append({
                            "id": f"E{len(events)+1:03d}",
                            "title": title_clean,
                            "date": date_str,
                            "venue": venue_str.split(",")[0].strip(),
                            "prizePool": "Rp 15.000.000 (Prize Pool)",
                            "entryFee": "Rp 150.000",
                            "contact": "0821-1502-3944 (Sekretariat POBSI)",
                            "status": status_val,
                            "description": f"Agenda turnamen sirkuit resmi POBSI Kabupaten Banjarnegara di {venue_str}.",
                            "poster": "images/event-poster.png"
                        })
            i += 3
        else:
            i += 1
            
    print(f"Parsed {len(events)} events from Kalender Event.")
    return events

def main():
    players = parse_rekap_hc()
    standings = parse_standings()
    events = parse_calendar()
    
    # Let's read existing documents from db.json so we preserve them
    db_json_path = r"d:\POBSI\api\data\db.json"
    documents = []
    if os.path.exists(db_json_path):
        try:
            with open(db_json_path, 'r', encoding='utf-8') as f:
                old_data = json.load(f)
                documents = old_data.get("documents", [])
        except Exception:
            pass
            
    if not documents:
        documents = [
            { "id": "D001", "title": "Surat Edaran Rapat Pengurus POBSI.pdf", "date": "20 Januari 2026", "fileSize": "528 KB", "fileType": "PDF" }
        ]
    else:
        # Update D001 size to match the actual file size if it's there
        for doc in documents:
            if doc["id"] == "D001":
                doc["title"] = "Surat Edaran Rapat Pengurus POBSI.pdf"
                doc["date"] = "20 Januari 2026"
                doc["fileSize"] = "515 KB"
                
    # 1. Save to api/data/db.json
    db_data = {
        "players": players,
        "standings": standings,
        "events": events,
        "documents": documents
    }
    
    # Ensure api/data directory exists
    os.makedirs(os.path.dirname(db_json_path), exist_ok=True)
    
    with open(db_json_path, 'w', encoding='utf-8') as f:
        json.dump(db_data, f, indent=2)
    print(f"Successfully saved all data to JSON backup: {db_json_path}")
    
    # 2. Save directly to SQLite database
    db_path = r"d:\POBSI\api\data\billiard.db"
    if not os.path.exists(db_path):
        print("SQLite database file not found yet. Running server will create it, but we can write directly.")
        
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create tables if not exist (ensuring TEXT handicap and player points!)
    cursor.execute("""CREATE TABLE IF NOT EXISTS players (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        club TEXT NOT NULL,
        handicap TEXT NOT NULL,
        status TEXT DEFAULT 'Aktif',
        points REAL DEFAULT 0.0,
        avatar TEXT,
        gender TEXT,
        age INTEGER,
        phone TEXT,
        address TEXT
    )""")
    
    cursor.execute("""CREATE TABLE IF NOT EXISTS standings (
        rank INTEGER,
        name TEXT PRIMARY KEY,
        club TEXT NOT NULL,
        handicap TEXT NOT NULL,
        points INTEGER NOT NULL,
        played INTEGER DEFAULT 0,
        won INTEGER DEFAULT 0,
        lost INTEGER DEFAULT 0,
        trend TEXT DEFAULT 'stable'
    )""")
    
    cursor.execute("""CREATE TABLE IF NOT EXISTS events (
        id TEXT PRIMARY KEY,
        title TEXT NOT NULL,
        date TEXT NOT NULL,
        venue TEXT NOT NULL,
        prizePool TEXT,
        entryFee TEXT,
        contact TEXT NOT NULL,
        status TEXT DEFAULT 'Daftar',
        description TEXT,
        poster TEXT
    )""")
    
    cursor.execute("""CREATE TABLE IF NOT EXISTS documents (
        id TEXT PRIMARY KEY,
        title TEXT NOT NULL,
        date TEXT NOT NULL,
        fileSize TEXT,
        fileType TEXT DEFAULT 'PDF'
    )""")
    
    # Clear existing data
    cursor.execute("DELETE FROM players")
    cursor.execute("DELETE FROM standings")
    cursor.execute("DELETE FROM events")
    cursor.execute("DELETE FROM documents")
    
    # Insert new data
    for p in players:
        cursor.execute(
            "INSERT INTO players (id, name, club, handicap, status, points, avatar, gender, age, phone, address) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (p["id"], p["name"], p["club"], p["handicap"], p["status"], p["points"], p["avatar"], p["gender"], p["age"], p["phone"], p["address"])
        )
        
    for s in standings:
        cursor.execute(
            "INSERT INTO standings (rank, name, club, handicap, points, played, won, lost, trend) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (s["rank"], s["name"], s["club"], s["handicap"], s["points"], s["played"], s["won"], s["lost"], s["trend"])
        )
        
    for e in events:
        cursor.execute(
            "INSERT INTO events (id, title, date, venue, prizePool, entryFee, contact, status, description, poster) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (e["id"], e["title"], e["date"], e["venue"], e["prizePool"], e["entryFee"], e["contact"], e["status"], e["description"], e["poster"])
        )
        
    for d in documents:
        cursor.execute(
            "INSERT INTO documents (id, title, date, fileSize, fileType) VALUES (?, ?, ?, ?, ?)",
            (d["id"], d["title"], d["date"], d["fileSize"], d["fileType"])
        )
        
    conn.commit()
    conn.close()
    print(f"Successfully synchronized all data to SQLite database {db_path}!")

if __name__ == "__main__":
    main()
